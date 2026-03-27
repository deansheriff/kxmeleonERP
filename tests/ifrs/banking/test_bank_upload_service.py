from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from unittest.mock import MagicMock

from app.services.finance.banking.bank_upload import BankUploadService, PaymentItem


def test_generate_zenith_format_produces_valid_excel(monkeypatch):
    """Zenith format produces an xlsx file with text-formatted account numbers."""
    from openpyxl import load_workbook

    service = BankUploadService(MagicMock())
    monkeypatch.setattr(service.bank_directory, "lookup_bank_code", lambda _name: "44")

    result = service.generate_upload(
        items=[
            PaymentItem(
                reference="PAY-001",
                beneficiary_name="Jane Doe",
                amount=Decimal("1500.50"),
                account_number="12345",
                bank_name="Some Bank",
                beneficiary_code="EMP001",
            )
        ],
        source_account_number="7890",
        payment_date=date(2026, 2, 1),
        bank_format="zenith",
    )

    assert result.filename == "bank_upload_zenith_20260201.xlsx"
    assert result.row_count == 1
    assert result.total_amount == Decimal("1500.50")
    assert result.errors == []
    assert "spreadsheetml" in result.content_type

    # Parse the Excel content and verify cell values
    wb = load_workbook(io.BytesIO(result.content))
    ws = wb.active

    # Header row — full descriptive text from Zenith template
    assert ws.cell(row=1, column=1).value.startswith("TRANSACTION REFERENCE NUMBER")
    assert ws.cell(row=1, column=6).value.startswith("BENEFICIARY ACCOUNT NUMBER")

    # Data row — account numbers zero-padded and text-formatted
    assert ws.cell(row=2, column=1).value == "PAY-001"
    assert ws.cell(row=2, column=2).value == "Jane Doe"
    assert ws.cell(row=2, column=3).value == 1500.50
    assert ws.cell(row=2, column=3).number_format == "0.00;[Red]0.00"
    assert ws.cell(row=2, column=5).value == "EMP001"
    assert ws.cell(row=2, column=6).value == "0000012345"  # zero-padded
    assert ws.cell(row=2, column=6).number_format == "@"  # text format
    assert ws.cell(row=2, column=7).value == "044"  # zero-padded bank code
    assert ws.cell(row=2, column=7).number_format == "@"
    assert ws.cell(row=2, column=8).value == "0000007890"  # zero-padded debit acct
    assert ws.cell(row=2, column=8).number_format == "@"


def test_generate_upload_reports_missing_bank_code(monkeypatch):
    service = BankUploadService(MagicMock())
    monkeypatch.setattr(service.bank_directory, "lookup_bank_code", lambda _name: None)

    result = service.generate_upload(
        items=[
            PaymentItem(
                reference="PAY-002",
                beneficiary_name="No Code Vendor",
                amount=Decimal("99.99"),
                account_number="0123456789",
                bank_name="Unknown Bank",
            )
        ],
        source_account_number="0000000001",
        payment_date=date(2026, 2, 1),
        bank_format="generic",
    )

    assert result.row_count == 1
    assert result.total_amount == Decimal("99.99")
    assert result.errors == ["Bank code not found for: No Code Vendor (Unknown Bank)"]


def test_resolve_bank_code_prefers_item_code_and_formats_digits(monkeypatch):
    service = BankUploadService(MagicMock())

    called = {"lookup": 0}

    def _lookup(_name: str):
        called["lookup"] += 1
        return "011"

    monkeypatch.setattr(service.bank_directory, "lookup_bank_code", _lookup)

    explicit_digit = service._resolve_bank_code(
        PaymentItem(
            reference="1",
            beneficiary_name="A",
            amount=Decimal("1"),
            account_number="1",
            bank_name="B",
            bank_code="57",
        )
    )
    explicit_text = service._resolve_bank_code(
        PaymentItem(
            reference="2",
            beneficiary_name="A",
            amount=Decimal("1"),
            account_number="1",
            bank_name="B",
            bank_code="ABC",
        )
    )
    looked_up = service._resolve_bank_code(
        PaymentItem(
            reference="3",
            beneficiary_name="A",
            amount=Decimal("1"),
            account_number="1",
            bank_name="B",
        )
    )

    assert explicit_digit == "057"
    assert explicit_text == "ABC"
    assert looked_up == "011"
    assert called["lookup"] == 1


def test_generate_access_uses_default_narration_when_missing(monkeypatch):
    service = BankUploadService(MagicMock())
    monkeypatch.setattr(service.bank_directory, "lookup_bank_code", lambda _name: "058")

    result = service.generate_upload(
        items=[
            PaymentItem(
                reference="PAY-003",
                beneficiary_name="John Smith",
                amount=Decimal("250.00"),
                account_number="22334455",
                bank_name="GTBank",
                narration=None,
            )
        ],
        source_account_number="0000000001",
        payment_date=date(2026, 2, 2),
        bank_format="access",
    )

    csv_text = result.content.decode("utf-8")
    assert "Payment to John Smith" in csv_text
    assert result.filename == "bank_upload_access_20260202.csv"
