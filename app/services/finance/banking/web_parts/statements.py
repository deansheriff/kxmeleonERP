"""BankingStatementWebService component."""

from __future__ import annotations

from app.services.finance.banking.web_parts.base import (
    Account,
    Any,
    BankAccount,
    BankAccountStatus,
    BankStatement,
    BankStatementImport,
    BankStatementLine,
    BytesIO,
    HTMLResponse,
    HTTPException,
    JSONResponse,
    JournalEntry,
    JournalEntryLine,
    RedirectResponse,
    Request,
    Response,
    SPREADSHEET_EXTENSIONS,
    Session,
    StringIO,
    UUID,
    UploadFile,
    ValidationError,
    WebAuthContext,
    _SOURCE_TYPE_LABELS,
    _account_view,
    _build_active_filters,
    _build_match_detail,
    _datetime,
    _parse_date,
    _parse_decimal,
    _parse_statement_status,
    _statement_line_view,
    _statement_view,
    apply_sort,
    bank_statement_service,
    base_context,
    build_active_filters,
    builtins,
    case,
    coerce_uuid,
    csv,
    date,
    func,
    logger,
    or_,
    org_context_service,
    re,
    resolve_payment_metadata_batch,
    select,
    spreadsheet_formats_label,
    templates,
)


class BankingStatementWebService:
    """Banking web service methods for statements."""

    @staticmethod
    def list_statements_context(
        db: Session,
        organization_id: str,
        account_id: str | None,
        status: str | None,
        start_date: str | None,
        end_date: str | None,
        page: int,
        limit: int = 50,
        sort: str | None = None,
        sort_dir: str | None = None,
        match_status: str | None = None,
        search: str | None = None,
    ) -> dict:
        """Build context for flat bank statement lines page.

        Shows bank account summary cards and imported BankStatementLine
        rows across all statements, with categorization and match status.
        """
        from app.models.finance.banking.bank_statement import CategorizationStatus

        org_id = coerce_uuid(organization_id)
        offset = (page - 1) * limit

        from_date = _parse_date(start_date)
        to_date = _parse_date(end_date)

        # ── Bank accounts (always shown as summary cards) ──
        accounts = list(
            db.scalars(
                select(BankAccount)
                .where(
                    BankAccount.organization_id == org_id,
                    BankAccount.status == BankAccountStatus.active,
                )
                .order_by(BankAccount.bank_name, BankAccount.account_number)
            ).all()
        )
        account_views = [_account_view(a) for a in accounts]

        # Map bank_account_id → BankAccount for enriching line views
        bank_map: dict[UUID, BankAccount] = {a.bank_account_id: a for a in accounts}

        # ── Base conditions for BankStatementLine query ──
        base_conditions: list[Any] = [
            BankStatement.organization_id == org_id,
        ]
        if account_id:
            base_conditions.append(
                BankStatement.bank_account_id == coerce_uuid(account_id)
            )
        if from_date:
            base_conditions.append(BankStatementLine.transaction_date >= from_date)
        if to_date:
            base_conditions.append(BankStatementLine.transaction_date <= to_date)
        if status:
            try:
                cat_status = CategorizationStatus(status)
                base_conditions.append(
                    BankStatementLine.categorization_status == cat_status
                )
            except ValueError:
                pass  # ignore invalid status filter
        if match_status == "matched":
            base_conditions.append(BankStatementLine.is_matched.is_(True))
        elif match_status == "unmatched":
            base_conditions.append(BankStatementLine.is_matched.is_(False))

        search_term = (search or "").strip()
        if search_term:
            like_pat = f"%{search_term}%"
            base_conditions.append(
                or_(
                    BankStatementLine.description.ilike(like_pat),
                    BankStatementLine.reference.ilike(like_pat),
                    BankStatementLine.payee_payer.ilike(like_pat),
                    BankStatementLine.bank_reference.ilike(like_pat),
                )
            )

        join_clause = BankStatementLine.statement_id == BankStatement.statement_id

        # Count
        count_stmt = (
            select(func.count(BankStatementLine.line_id))
            .join(BankStatement, join_clause)
            .where(*base_conditions)
        )
        total_count = db.scalar(count_stmt) or 0

        # Aggregates for stat cards
        agg_stmt = (
            select(
                func.count(BankStatementLine.line_id).label("total"),
                func.count(
                    case(
                        (
                            BankStatementLine.categorization_status.is_(None),
                            BankStatementLine.line_id,
                        ),
                    )
                ).label("uncategorized"),
                func.count(
                    case(
                        (
                            BankStatementLine.categorization_status.in_(
                                [
                                    "SUGGESTED",
                                    "FLAGGED",
                                ]
                            ),
                            BankStatementLine.line_id,
                        ),
                    )
                ).label("suggested"),
                func.count(
                    case(
                        (
                            BankStatementLine.is_matched.is_(True),
                            BankStatementLine.line_id,
                        ),
                    )
                ).label("matched"),
                func.count(
                    case(
                        (
                            BankStatementLine.is_matched.is_(False),
                            BankStatementLine.line_id,
                        ),
                    )
                ).label("unmatched"),
            )
            .join(BankStatement, join_clause)
            .where(*base_conditions)
        )
        agg_row = db.execute(agg_stmt).one()
        total_lines = agg_row.total or 0
        uncategorized_count = agg_row.uncategorized or 0
        suggested_count = agg_row.suggested or 0
        matched_count = agg_row.matched or 0
        unmatched_count = agg_row.unmatched or 0

        # Fetch paginated statement lines
        txn_sort_map: dict[str, Any] = {
            "transaction_date": BankStatementLine.transaction_date,
            "amount": BankStatementLine.amount,
            "description": BankStatementLine.description,
        }
        txn_stmt = (
            select(BankStatementLine, BankStatement)
            .join(BankStatement, join_clause)
            .where(*base_conditions)
        )
        txn_stmt = apply_sort(
            txn_stmt,
            sort,
            sort_dir,
            txn_sort_map,
            default=BankStatementLine.transaction_date.desc(),
        )
        rows = db.execute(txn_stmt.limit(limit).offset(offset)).all()

        # Build line view dicts enriched with bank account info
        transactions: list[dict[str, Any]] = []
        suggested_account_ids: set[UUID] = set()
        matched_jl_ids: set[UUID] = set()
        for line, stmt in rows:
            bank_acct = bank_map.get(stmt.bank_account_id)
            currency = (
                stmt.currency_code
                or (bank_acct.currency_code if bank_acct else None)
                or org_context_service.get_functional_currency(db, organization_id)
            )
            txn = _statement_line_view(line, currency)
            txn["statement_id"] = str(stmt.statement_id)
            txn["statement_number"] = stmt.statement_number or ""
            txn["bank_name"] = bank_acct.bank_name if bank_acct else ""
            txn["account_number"] = bank_acct.account_number if bank_acct else ""
            txn["bank_account_id"] = str(stmt.bank_account_id)
            transactions.append(txn)
            if line.suggested_account_id:
                suggested_account_ids.add(line.suggested_account_id)
            if line.matched_journal_line_id:
                matched_jl_ids.add(line.matched_journal_line_id)

        # Batch-resolve matched journal entry info
        match_detail_map: dict[str, dict[str, str]] = {}
        if matched_jl_ids:
            jl_rows = db.execute(
                select(JournalEntryLine, JournalEntry)
                .join(
                    JournalEntry,
                    JournalEntryLine.journal_entry_id == JournalEntry.journal_entry_id,
                )
                .where(JournalEntryLine.line_id.in_(list(matched_jl_ids)))
            ).all()

            # Batch-resolve payment metadata for source labels
            md_pairs: list[tuple[str | None, UUID | None]] = [
                (
                    getattr(je, "source_document_type", None),
                    getattr(je, "source_document_id", None),
                )
                for _jl, je in jl_rows
            ]
            md_map = resolve_payment_metadata_batch(db, md_pairs)

            for jl, je in jl_rows:
                doc_id = getattr(je, "source_document_id", None)
                meta = md_map.get(doc_id) if doc_id else None
                src_type = getattr(je, "source_document_type", None) or ""
                source_label = _SOURCE_TYPE_LABELS.get(src_type, "Journal")
                counterparty = ""
                if meta and meta.counterparty_name:
                    counterparty = meta.counterparty_name
                match_detail_map[str(jl.line_id)] = {
                    "journal_entry_id": str(je.journal_entry_id),
                    "journal_number": je.journal_number or "",
                    "source_label": source_label,
                    "counterparty": counterparty,
                }

        # Enrich transactions with match details
        for txn in transactions:
            jl_id = txn.get("matched_journal_line_id")
            detail = match_detail_map.get(jl_id or "") if jl_id else None
            if detail:
                txn["match_journal_entry_id"] = detail["journal_entry_id"]
                txn["match_journal_number"] = detail["journal_number"]
                txn["match_source_label"] = detail["source_label"]
                txn["match_counterparty"] = detail["counterparty"]

        # Build account name map for suggested accounts
        account_map: dict[str, str] = {}
        if suggested_account_ids:
            gl_accounts = db.scalars(
                select(Account).where(
                    Account.organization_id == org_id,
                    Account.account_id.in_(list(suggested_account_ids)),
                )
            ).all()
            account_map = {
                str(a.account_id): f"{a.account_code} - {a.account_name}"
                for a in gl_accounts
            }

        # Always show the Category column so the workflow is discoverable
        has_category_data = True

        total_pages = max(1, (total_count + limit - 1) // limit)

        # Status filter labels
        cat_status_labels: dict[str, str] = {
            "SUGGESTED": "Suggested",
            "ACCEPTED": "Accepted",
            "REJECTED": "Rejected",
            "AUTO_APPLIED": "Auto-applied",
            "FLAGGED": "Flagged",
        }
        match_status_labels: dict[str, str] = {
            "matched": "Matched",
            "unmatched": "Unmatched",
        }

        active_filters = _build_active_filters(
            account_id=account_id,
            accounts=account_views,
            status=status,
            start_date=start_date,
            end_date=end_date,
            status_labels=cat_status_labels,
        )
        if match_status and match_status in match_status_labels:
            active_filters.append(
                {
                    "name": "match_status",
                    "value": match_status,
                    "display_value": match_status_labels[match_status],
                }
            )
        if search_term:
            active_filters.append(
                {
                    "name": "search",
                    "value": search_term,
                    "display_value": f'Search: "{search_term}"',
                }
            )

        return {
            "transactions": transactions,
            "accounts": account_views,
            "account_id": account_id,
            "status": status,
            "match_status": match_status,
            "search": search_term,
            "sort": sort,
            "sort_dir": sort_dir,
            "start_date": start_date,
            "end_date": end_date,
            "page": page,
            "limit": limit,
            "offset": offset,
            "total_count": total_count,
            "total_pages": total_pages,
            "total_lines": total_lines,
            "uncategorized_count": uncategorized_count,
            "suggested_count": suggested_count,
            "matched_count": matched_count,
            "unmatched_count": unmatched_count,
            "has_category_data": has_category_data,
            "account_map": account_map,
            "active_filters": active_filters,
        }

    @staticmethod
    def statement_import_context(
        db: Session,
        organization_id: str,
    ) -> dict:
        org_id = coerce_uuid(organization_id)
        accounts = db.scalars(
            select(BankAccount)
            .where(BankAccount.organization_id == org_id)
            .order_by(BankAccount.bank_name, BankAccount.account_number)
        ).all()
        # Build JSON-safe account list for Alpine.js tojson serialization.
        accounts_json = [
            {
                "bank_account_id": str(a.bank_account_id),
                "bank_name": a.bank_name or "",
                "account_number": a.account_number or "",
            }
            for a in accounts
        ]
        # Build alias map from unified registry for client-side header matching
        from app.services.finance.banking.bank_statement import (
            BankStatementService,
        )
        from app.services.finance.import_export.base import build_alias_map

        alias_map = build_alias_map(BankStatementService._BANK_FIELD_TYPES)
        return {"accounts": accounts_json, "alias_map": alias_map}

    @staticmethod
    def statement_detail_context(
        db: Session,
        organization_id: str,
        statement_id: str,
        page: int = 1,
        limit: int = 50,
    ) -> dict:
        org_id = coerce_uuid(organization_id)
        statement = db.get(BankStatement, coerce_uuid(statement_id))
        if not statement or statement.organization_id != org_id:
            return {"statement": None, "lines": [], "account_map": {}}

        currency = statement.currency_code
        total_count = (
            db.scalar(
                select(func.count(BankStatementLine.line_id)).where(
                    BankStatementLine.statement_id == statement.statement_id
                )
            )
            or 0
        )
        total_pages = max(1, (total_count + limit - 1) // limit)
        offset = (page - 1) * limit
        paged_lines = list(
            db.scalars(
                select(BankStatementLine)
                .where(BankStatementLine.statement_id == statement.statement_id)
                .order_by(BankStatementLine.transaction_date)
                .offset(offset)
                .limit(limit)
            ).all()
        )
        lines = [_statement_line_view(line, currency) for line in paged_lines]

        # Build account name lookup for suggested accounts
        account_ids = [
            line.suggested_account_id
            for line in paged_lines
            if line.suggested_account_id
        ]
        account_map: dict[str, str] = {}
        if account_ids:
            accounts = db.scalars(
                select(Account).where(
                    Account.organization_id == org_id,
                    Account.account_id.in_(account_ids),
                )
            ).all()
            account_map = {
                str(a.account_id): f"{a.account_code} - {a.account_name}"
                for a in accounts
            }

        # Categorization summary counts (SQL aggregation, not in-memory)
        from app.models.finance.banking.bank_statement import CategorizationStatus

        cat_rows = db.execute(
            select(
                BankStatementLine.categorization_status,
                func.count(BankStatementLine.line_id),
            )
            .where(BankStatementLine.statement_id == statement.statement_id)
            .group_by(BankStatementLine.categorization_status)
        ).all()
        cat_summary = {
            "suggested": 0,
            "accepted": 0,
            "rejected": 0,
            "auto_applied": 0,
            "flagged": 0,
        }
        _cat_key_map = {
            CategorizationStatus.SUGGESTED: "suggested",
            CategorizationStatus.ACCEPTED: "accepted",
            CategorizationStatus.REJECTED: "rejected",
            CategorizationStatus.AUTO_APPLIED: "auto_applied",
            CategorizationStatus.FLAGGED: "flagged",
        }
        for cat_status, cnt in cat_rows:
            key = _cat_key_map.get(cat_status)
            if key:
                cat_summary[key] = cnt

        # GL transaction match suggestions for unmatched lines
        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
        )

        recon_svc = BankReconciliationService()
        match_suggestions_raw = recon_svc.get_statement_match_suggestions(
            db, org_id, statement.statement_id
        )

        # Serialize to JSON-safe dict keyed by string line_id
        match_suggestions: dict[str, dict] = {}
        visible_line_ids = {str(line.line_id) for line in paged_lines}
        for line_id, suggestion in match_suggestions_raw.items():
            line_id_str = str(line_id)
            if line_id_str not in visible_line_ids:
                continue
            match_suggestions[line_id_str] = {
                "journal_line_id": str(suggestion.journal_line_id),
                "confidence": suggestion.confidence,
                "counterparty_name": suggestion.counterparty_name,
                "payment_number": suggestion.payment_number,
                "source_url": suggestion.source_url,
                "amount_matched": suggestion.amount_matched,
            }

        # GL candidates are now lazy-loaded per line via the scored
        # candidates API endpoint — no longer sent at page load.

        # Resolve source URLs for matched lines and build line_amounts map
        from app.services.finance.banking.bank_reconciliation import (
            _build_source_url,
        )

        matched_jl_ids = [
            line.matched_journal_line_id
            for line in paged_lines
            if line.is_matched and line.matched_journal_line_id
        ]
        matched_source_urls: dict[str, str] = {}
        # Map journal_line_id → JournalEntry for metadata resolution
        jl_entry_map: dict[str, JournalEntry] = {}
        if matched_jl_ids:
            jl_rows = (
                db.execute(
                    select(JournalEntryLine)
                    .join(JournalEntry)
                    .where(JournalEntryLine.line_id.in_(matched_jl_ids))
                )
                .scalars()
                .all()
            )
            for jl in jl_rows:
                entry = getattr(jl, "journal_entry", None) or getattr(jl, "entry", None)
                if entry:
                    url = _build_source_url(
                        getattr(entry, "source_document_type", None),
                        getattr(entry, "source_document_id", None),
                        getattr(entry, "entry_id", None),
                    )
                    matched_source_urls[str(jl.line_id)] = url
                    jl_entry_map[str(jl.line_id)] = entry

        # Batch-resolve payment metadata for matched lines
        metadata_pairs: list[tuple[str | None, UUID | None]] = [
            (
                getattr(e, "source_document_type", None),
                getattr(e, "source_document_id", None),
            )
            for e in jl_entry_map.values()
        ]
        metadata_by_doc_id = resolve_payment_metadata_batch(db, metadata_pairs)

        # Build match_details keyed by statement line_id
        # Map statement line → matched journal line id for lookup
        stmt_line_to_jl: dict[str, str] = {}
        for line in paged_lines:
            if line.is_matched and line.matched_journal_line_id:
                stmt_line_to_jl[str(line.line_id)] = str(line.matched_journal_line_id)

        match_details: dict[str, dict[str, str]] = {}
        for stmt_lid, jl_id in stmt_line_to_jl.items():
            entry = jl_entry_map.get(jl_id)
            if not entry:
                continue
            src_doc_id = getattr(entry, "source_document_id", None)
            meta = metadata_by_doc_id.get(src_doc_id) if src_doc_id else None
            url = matched_source_urls.get(jl_id, "")
            match_details[stmt_lid] = _build_match_detail(db, entry, url, metadata=meta)

        # Merge matched_source_url into line views
        line_amounts: dict[str, float] = {}
        for lv in lines:
            lid = str(lv["line_id"])
            line_amounts[lid] = lv["raw_amount"]
            matched_journal_line_id = lv.get("matched_journal_line_id")
            if lv["is_matched"] and matched_journal_line_id:
                lv["matched_source_url"] = matched_source_urls.get(
                    str(matched_journal_line_id), ""
                )
            else:
                lv["matched_source_url"] = ""

        # Build line_details for modal context card (issue #4)
        # Keyed by string line_id with essential info for visual comparison
        line_details: dict[str, dict] = {}
        for lv in lines:
            lid = str(lv["line_id"])
            line_details[lid] = {
                "date": lv["transaction_date"],
                "description": lv["description"] or "",
                "payee": lv["payee_payer"] or "",
                "amount": lv["amount"],
                "raw_amount": lv["raw_amount"],
                "is_credit": lv["transaction_type"] == "credit",
            }

        # Check if any lines have balance/category data (issue #16/#17)
        has_balance_data = any(line.running_balance is not None for line in paged_lines)
        has_category_data = any(
            line.categorization_status is not None for line in paged_lines
        )

        # GL accounts for "Create Journal & Match" feature
        from app.models.finance.gl.account import Account as GLAccount

        gl_accounts_raw = list(
            db.scalars(
                select(GLAccount)
                .where(
                    GLAccount.organization_id == org_id,
                    GLAccount.is_active == True,  # noqa: E712
                )
                .order_by(GLAccount.account_code)
            ).all()
        )
        gl_accounts = [
            {
                "id": str(a.account_id),
                "code": a.account_code,
                "name": a.account_name,
                "label": f"{a.account_code} - {a.account_name}",
            }
            for a in gl_accounts_raw
        ]
        # Bank account's GL account id for filtering
        bank_gl_account_id = (
            str(statement.bank_account.gl_account_id)
            if statement.bank_account
            and getattr(statement.bank_account, "gl_account_id", None)
            else ""
        )

        return {
            "statement": _statement_view(statement),
            "lines": lines,
            "page": page,
            "limit": limit,
            "offset": offset,
            "total_count": total_count,
            "total_pages": total_pages,
            "account_map": account_map,
            "categorization_summary": cat_summary,
            "match_suggestions": match_suggestions,
            "match_details": match_details,
            "line_amounts": line_amounts,
            "line_details": line_details,
            "statement_currency": currency,
            "has_balance_data": has_balance_data,
            "has_category_data": has_category_data,
            "gl_accounts": gl_accounts,
            "bank_gl_account_id": bank_gl_account_id,
        }

    def list_statements_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        account_id: str | None,
        status: str | None,
        start_date: str | None,
        end_date: str | None,
        page: int,
        limit: int = 50,
        sort: str | None = None,
        sort_dir: str | None = None,
        match_status: str | None = None,
        search: str | None = None,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Bank Transactions", "banking", db=db)
        context.update(
            self.list_statements_context(
                db,
                str(auth.organization_id),
                account_id=account_id,
                status=status,
                start_date=start_date,
                end_date=end_date,
                page=page,
                limit=limit,
                sort=sort,
                sort_dir=sort_dir,
                match_status=match_status,
                search=search,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/statements.html", context
        )

    @staticmethod
    def list_statement_imports_context(
        db: Session,
        organization_id: str,
        account_id: str | None,
        status: str | None,
        start_date: str | None,
        end_date: str | None,
        page: int,
        limit: int = 25,
        sort: str | None = None,
        sort_dir: str | None = None,
        search: str | None = None,
    ) -> dict:
        """Build context for statement imports list (header-level view).

        Shows imported BankStatement headers grouped by bank account,
        each linking to the per-statement detail/matching page.
        """
        org_id = coerce_uuid(organization_id)
        offset = (page - 1) * limit

        from_date = _parse_date(start_date)
        to_date = _parse_date(end_date)

        # ── Bank accounts for filter dropdown ──
        accounts = list(
            db.scalars(
                select(BankAccount)
                .where(
                    BankAccount.organization_id == org_id,
                    BankAccount.status == BankAccountStatus.active,
                )
                .order_by(BankAccount.bank_name, BankAccount.account_number)
            ).all()
        )

        # ── Statement headers query ──
        stmt = (
            select(BankStatement)
            .join(
                BankAccount,
                BankStatement.bank_account_id == BankAccount.bank_account_id,
            )
            .where(BankStatement.organization_id == org_id)
        )

        if account_id:
            stmt = stmt.where(BankStatement.bank_account_id == coerce_uuid(account_id))
        parsed_status = _parse_statement_status(status)
        if parsed_status:
            stmt = stmt.where(BankStatement.status == parsed_status)
        if from_date:
            stmt = stmt.where(BankStatement.statement_date >= from_date)
        if to_date:
            stmt = stmt.where(BankStatement.statement_date <= to_date)
        if search:
            stmt = stmt.where(BankStatement.statement_number.ilike(f"%{search}%"))

        # ── Sorting ──
        sort_col = sort or "statement_date"
        col_map: dict[str, Any] = {
            "statement_date": BankStatement.statement_date,
            "statement_number": BankStatement.statement_number,
            "total_lines": BankStatement.total_lines,
        }
        order_col = col_map.get(sort_col, BankStatement.statement_date)
        if (sort_dir or "desc").lower() == "asc":
            stmt = stmt.order_by(order_col.asc())
        else:
            stmt = stmt.order_by(order_col.desc())

        # ── Count + paginate ──
        count_stmt = select(func.count()).select_from(stmt.subquery())
        total_count: int = db.scalar(count_stmt) or 0
        total_pages = max(1, (total_count + limit - 1) // limit)

        statements = list(db.scalars(stmt.offset(offset).limit(limit)).all())

        # ── Real matched counts from BankStatementLine ──
        # The BankStatement.matched_lines header column is not reliably updated,
        # so we compute actual counts from the line-level is_matched flag.
        stmt_ids = [s.statement_id for s in statements]
        line_counts: dict[UUID, dict[str, int]] = {}
        if stmt_ids:
            count_rows = db.execute(
                select(
                    BankStatementLine.statement_id,
                    func.count().label("total"),
                    func.count()
                    .filter(BankStatementLine.is_matched.is_(True))
                    .label("matched"),
                )
                .where(BankStatementLine.statement_id.in_(stmt_ids))
                .group_by(BankStatementLine.statement_id)
            ).all()
            for row in count_rows:
                line_counts[row.statement_id] = {
                    "total": row.total,
                    "matched": row.matched,
                }

        total_lines = sum(lc["total"] for lc in line_counts.values())
        matched_lines = sum(lc["matched"] for lc in line_counts.values())

        # ── Active filters ──
        active_filters = build_active_filters(
            params={
                "account_id": account_id,
                "status": status,
                "search": search,
                "start_date": start_date,
                "end_date": end_date,
            },
            labels={
                "status": "Status",
                "search": "Search",
                "start_date": "From",
                "end_date": "To",
            },
            options={
                "account_id": {
                    str(a.bank_account_id): f"{a.bank_name} - {a.account_number}"
                    for a in accounts
                },
                "status": {
                    "imported": "Imported",
                    "processing": "Processing",
                    "closed": "Reconciled",
                },
            },
        )

        # ── Build statement views with real matched counts ──
        statement_views = []
        for s in statements:
            sv = _statement_view(s)
            lc = line_counts.get(s.statement_id, {"total": 0, "matched": 0})
            sv["total_lines"] = lc["total"]
            sv["matched_lines"] = lc["matched"]
            sv["unmatched_lines"] = lc["total"] - lc["matched"]
            statement_views.append(sv)

        return {
            "statements": statement_views,
            "accounts": [_account_view(a) for a in accounts],
            "account_id": account_id or "",
            "status": status or "",
            "search": search or "",
            "start_date": start_date or "",
            "end_date": end_date or "",
            "sort": sort_col,
            "sort_dir": sort_dir or "desc",
            "page": page,
            "limit": limit,
            "offset": offset,
            "total_count": total_count,
            "total_pages": total_pages,
            "total_lines": total_lines,
            "matched_lines": matched_lines,
            "unmatched_lines": total_lines - matched_lines,
            "active_filters": active_filters,
        }

    def list_statement_imports_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        account_id: str | None,
        status: str | None,
        start_date: str | None,
        end_date: str | None,
        page: int,
        limit: int = 25,
        sort: str | None = None,
        sort_dir: str | None = None,
        search: str | None = None,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Imported Statements", "banking", db=db)
        context.update(
            self.list_statement_imports_context(
                db,
                str(auth.organization_id),
                account_id=account_id,
                status=status,
                start_date=start_date,
                end_date=end_date,
                page=page,
                limit=limit,
                sort=sort,
                sort_dir=sort_dir,
                search=search,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/statement_imports.html", context
        )

    def statement_import_form_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        form_data: dict | None = None,
        errors: list[str] | None = None,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Import Bank Statement", "banking", db=db)
        context.update(self.statement_import_context(db, str(auth.organization_id)))
        form_payload = form_data or {}
        if not form_payload:
            if request.query_params.get("account_id"):
                form_payload["bank_account_id"] = request.query_params.get("account_id")
            form_payload.setdefault("statement_date", date.today().isoformat())
        context["form_data"] = form_payload
        context["form_errors"] = errors or []
        return templates.TemplateResponse(
            request, "finance/banking/statement_import.html", context
        )

    async def statement_import_submit_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ):
        form = getattr(request.state, "csrf_form", None)
        if form is None:
            form = await request.form()
        form_data = dict(form)
        csv_format = form_data.get("csv_format") or None
        errors: list[str] = []
        org_date_fmt = self._resolve_org_date_format(db, auth.organization_id)

        # Parse lines from file or manual entry.
        lines_data: list[dict] = []
        upload = form.get("statement_file")
        upload_file = upload if isinstance(upload, UploadFile) else None
        has_upload = bool(upload_file and upload_file.filename)
        upload_ext = ""
        column_map = self._parse_column_map(form)
        mapped_lines_data, manual_errors = self._parse_manual_lines(form)
        # When a file is uploaded, ignore manual line fields from the same form
        # submission to avoid stale hidden inputs overriding uploaded data.
        if has_upload:
            mapped_lines_data = []
            manual_errors = []
        if mapped_lines_data:
            lines_data = self._normalize_mapped_lines(
                mapped_lines_data, org_date_fmt=org_date_fmt
            )
            errors.extend(manual_errors)
        if has_upload and upload_file is not None:
            # CSRF middleware parses form data first, which can advance the file pointer.
            try:
                await upload_file.seek(0)
            except (OSError, AttributeError):
                try:
                    upload_file.file.seek(0)
                except (OSError, AttributeError):
                    logger.warning("Could not seek upload file to start")
            filename = upload_file.filename or ""
            lowered = filename.lower()
            upload_ext = (
                ".csv"
                if lowered.endswith(".csv")
                else ".xls"
                if lowered.endswith(".xls")
                else ".xlsx"
                if lowered.endswith(".xlsx")
                else ".xlsm"
                if lowered.endswith(".xlsm")
                else ""
            )
            _ALLOWED_CONTENT_TYPES = {
                "text/csv",
                "application/csv",
                "application/vnd.ms-excel",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel.sheet.macroEnabled.12",
                "application/octet-stream",  # many browsers send this for .xls/.xlsx
            }
            if not lowered.endswith(SPREADSHEET_EXTENSIONS):
                errors.append(
                    f"Supported statement files: {spreadsheet_formats_label()}."
                )
            elif (
                upload_file.content_type
                and upload_file.content_type not in _ALLOWED_CONTENT_TYPES
            ):
                errors.append(
                    f"Invalid file type '{upload_file.content_type}'. "
                    f"Supported: {spreadsheet_formats_label()}."
                )
            else:
                # Limit upload size to avoid memory blowups.
                max_bytes = 10 * 1024 * 1024  # 10 MiB
                content = await upload_file.read(max_bytes + 1)
                if len(content) > max_bytes:
                    errors.append(
                        "Uploaded file is too large (max 10 MB). Please upload a smaller file."
                    )
                    content = b""
                elif not content:
                    # Some middleware (e.g., CSRF) may have consumed the file stream.
                    # Fall back to reading from the underlying file handle.
                    try:
                        upload_file.file.seek(0)
                        content = upload_file.file.read(max_bytes + 1)
                        if len(content) > max_bytes:
                            errors.append(
                                "Uploaded file is too large (max 10 MB). Please upload a smaller file."
                            )
                            content = b""
                    except OSError:
                        content = b""
                if content:
                    # Avoid logging file contents (may contain PII).
                    logger.info(
                        "Statement import file read: filename=%s bytes=%s content_type=%s",
                        upload_file.filename,
                        len(content),
                        upload_file.content_type,
                    )
                if not content:
                    logger.warning(
                        "Statement import upload empty after read: filename=%s content_type=%s",
                        upload_file.filename,
                        upload_file.content_type,
                    )
                    errors.append(
                        "Uploaded file appears empty. Please re-select the file and try again."
                    )
                else:
                    if column_map:
                        try:
                            _, source_rows, _total_rows = self._preview_upload_content(
                                content, lowered, sample_limit=None
                            )
                        except ValueError as exc:
                            errors.append(str(exc))
                            source_rows = []
                        mapped_rows = self._map_rows_with_column_map(
                            source_rows, column_map
                        )
                        lines_data = self._normalize_mapped_lines(
                            mapped_rows, org_date_fmt=org_date_fmt
                        )
                        if not lines_data:
                            errors.append(
                                "No rows found after applying column mapping. Please review your selected columns."
                            )
                    elif lowered.endswith(".csv"):
                        rows, parse_errors = bank_statement_service.parse_csv_rows(
                            content, csv_format, date_format=org_date_fmt
                        )
                        lines_data = rows
                        errors.extend(parse_errors)
                    elif lowered.endswith(".xls"):
                        rows, parse_errors = bank_statement_service.parse_xls_rows(
                            content, csv_format, date_format=org_date_fmt
                        )
                        lines_data = rows
                        errors.extend(parse_errors)
                    else:
                        rows, parse_errors = bank_statement_service.parse_xlsx_rows(
                            content, csv_format, date_format=org_date_fmt
                        )
                        lines_data = rows
                        errors.extend(parse_errors)
                    if not lines_data and not errors:
                        logger.warning(
                            "Statement import parsed zero rows: filename=%s csv_format=%s",
                            upload_file.filename,
                            csv_format,
                        )
        elif not lines_data:
            lines_data = mapped_lines_data
            errors.extend(manual_errors)

        if not lines_data and not errors:
            errors.append(
                "Please upload a CSV/Excel file or add at least one transaction."
            )

        payload_data = {
            "bank_account_id": form_data.get("bank_account_id"),
            "statement_number": form_data.get("statement_number") or None,
            "statement_date": form_data.get("statement_date") or None,
            "period_start": form_data.get("period_start"),
            "period_end": form_data.get("period_end"),
            "opening_balance": form_data.get("opening_balance") or None,
            "closing_balance": form_data.get("closing_balance") or None,
            "import_source": (
                "csv"
                if upload_ext == ".csv"
                else "excel"
                if upload_ext in {".xls", ".xlsx", ".xlsm"}
                else "manual"
            ),
            "import_filename": upload_file.filename if upload_file else None,
            "lines": lines_data,
        }

        payload = None
        if not errors:
            try:
                payload = BankStatementImport.model_validate(payload_data)
            except ValidationError as exc:
                errors.extend(self._format_validation_errors(exc))

        if errors or payload is None:
            # Preserve select fields for the form.
            preserved = {
                "bank_account_id": form_data.get("bank_account_id"),
                "statement_number": form_data.get("statement_number"),
                "statement_date": form_data.get("statement_date"),
                "period_start": form_data.get("period_start"),
                "period_end": form_data.get("period_end"),
                "opening_balance": form_data.get("opening_balance"),
                "closing_balance": form_data.get("closing_balance"),
                "csv_format": csv_format,
            }
            return self.statement_import_form_response(
                request, auth, db, form_data=preserved, errors=errors
            )

        line_inputs, line_errors = bank_statement_service.build_line_inputs(
            payload.lines
        )
        if line_errors:
            preserved = {
                "bank_account_id": form_data.get("bank_account_id"),
                "statement_number": form_data.get("statement_number"),
                "statement_date": form_data.get("statement_date"),
                "period_start": form_data.get("period_start"),
                "period_end": form_data.get("period_end"),
                "opening_balance": form_data.get("opening_balance"),
                "closing_balance": form_data.get("closing_balance"),
                "csv_format": csv_format,
            }
            return self.statement_import_form_response(
                request, auth, db, form_data=preserved, errors=line_errors
            )

        if auth.organization_id is None:
            raise HTTPException(status_code=400, detail="Organization is required")

        result = bank_statement_service.import_statement(
            db=db,
            organization_id=auth.organization_id,
            bank_account_id=payload.bank_account_id,
            statement_number=payload.statement_number,
            statement_date=payload.statement_date,
            period_start=payload.period_start,
            period_end=payload.period_end,
            opening_balance=payload.opening_balance,
            closing_balance=payload.closing_balance,
            lines=line_inputs,
            import_source=payload.import_source,
            import_filename=payload.import_filename,
            imported_by=auth.user_id,
        )
        db.flush()
        db.commit()
        redirect_url = (
            f"/finance/banking/statements/{result.statement.statement_id}"
            f"?success=Statement+imported+successfully"
            f"+({result.lines_imported}+lines)"
        )
        if result.auto_matched > 0:
            redirect_url += f"&auto_matched={result.auto_matched}"
        return RedirectResponse(url=redirect_url, status_code=303)

    async def statement_import_preview_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ) -> JSONResponse:
        form = getattr(request.state, "csrf_form", None)
        if form is None:
            form = await request.form()

        upload = form.get("statement_file")
        upload_file = upload if isinstance(upload, UploadFile) else None
        if not upload_file or not upload_file.filename:
            return JSONResponse(
                status_code=400,
                content={"detail": "Please choose a file to preview."},
            )

        filename = upload_file.filename or ""
        lowered = filename.lower()
        if not lowered.endswith(SPREADSHEET_EXTENSIONS):
            return JSONResponse(
                status_code=400,
                content={
                    "detail": (
                        f"Supported statement files: {spreadsheet_formats_label()}."
                    )
                },
            )

        try:
            await upload_file.seek(0)
        except Exception:
            try:
                upload_file.file.seek(0)
            except Exception:
                logger.exception("Ignored exception")

        max_bytes = 10 * 1024 * 1024  # 10 MiB
        content = await upload_file.read(max_bytes + 1)
        if len(content) > max_bytes:
            return JSONResponse(
                status_code=400,
                content={
                    "detail": "Uploaded file is too large (max 10 MB). Please upload a smaller file."
                },
            )
        if not content:
            try:
                upload_file.file.seek(0)
                content = upload_file.file.read(max_bytes + 1)
            except Exception:
                content = b""
        if not content:
            return JSONResponse(
                status_code=400,
                content={"detail": "Uploaded file appears empty."},
            )

        try:
            headers, sample_rows, total_rows = self._preview_upload_content(
                content, lowered
            )
        except ValueError as exc:
            return JSONResponse(status_code=400, content={"detail": str(exc)})

        return JSONResponse(
            {
                "detected_columns": headers,
                "sample_data": sample_rows,
                "total_rows": total_rows,
            }
        )

    @staticmethod
    def _preview_upload_content(
        content: bytes,
        lowered_filename: str,
        sample_limit: int | None = 5,
    ) -> tuple[list[str], list[dict[str, str]], int]:
        if lowered_filename.endswith(".csv"):
            try:
                text = content.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = content.decode("utf-8", errors="replace")

            delimiter = ","
            try:
                header_line = next(
                    (line for line in text.splitlines() if line.strip()), ""
                )
                candidates = [",", "\t", ";", "|"]
                counts = {d: header_line.count(d) for d in candidates}
                best = max(counts, key=counts.__getitem__)
                if counts.get(best, 0) > 0:
                    delimiter = best
                else:
                    sniff = csv.Sniffer().sniff(text[:2048])
                    if sniff and getattr(sniff, "delimiter", None):
                        delimiter = sniff.delimiter
            except Exception:
                delimiter = ","

            reader = csv.DictReader(StringIO(text), delimiter=delimiter)
            if not reader.fieldnames:
                raise ValueError("CSV file must include a header row.")
            headers = [str(h).strip() for h in reader.fieldnames if h is not None]
            # Keep DictReader row keys aligned with the trimmed headers shown in UI.
            reader.fieldnames = headers
            rows: list[dict[str, str]] = []
            total_rows = 0
            for row in reader:
                if not any(str(v).strip() for v in row.values() if v is not None):
                    continue
                total_rows += 1
                sample = {
                    h: ("" if row.get(h) is None else str(row.get(h))) for h in headers
                }
                if sample_limit is None or len(rows) < sample_limit:
                    rows.append(sample)
            return headers, rows, total_rows

        if lowered_filename.endswith(".xlsx") or lowered_filename.endswith(".xlsm"):
            from openpyxl import load_workbook

            workbook = load_workbook(
                filename=BytesIO(content), read_only=True, data_only=True
            )
            try:
                sheet = workbook.active
                rows_iter = sheet.iter_rows(values_only=True)
                try:
                    header_values = next(rows_iter)
                except StopIteration:
                    raise ValueError("Excel file must include a header row.") from None
                if not header_values:
                    raise ValueError("Excel file must include a header row.")
                headers = [
                    str(h).strip() if h is not None else "" for h in header_values
                ]
                headers = [h for h in headers if h]
                if not headers:
                    raise ValueError("Excel file must include a header row.")
                xlsx_rows: list[dict[str, str]] = []
                total_rows = 0
                for values in rows_iter:
                    if not values or not any(
                        value is not None and str(value).strip() for value in values
                    ):
                        continue
                    total_rows += 1
                    xlsx_row: dict[str, str] = {}
                    for i, header in enumerate(headers):
                        value = values[i] if i < len(values) else ""
                        if isinstance(value, _datetime):
                            xlsx_row[header] = value.strftime("%Y-%m-%d")
                        elif isinstance(value, date):
                            xlsx_row[header] = value.isoformat()
                        elif value is None:
                            xlsx_row[header] = ""
                        else:
                            xlsx_row[header] = str(value)
                    if sample_limit is None or len(xlsx_rows) < sample_limit:
                        xlsx_rows.append(xlsx_row)
                return headers, xlsx_rows, total_rows
            finally:
                workbook.close()

        if lowered_filename.endswith(".xls"):
            try:
                import xlrd
            except builtins.ImportError as exc:
                raise ValueError(
                    "XLS preview requires xlrd. Please install xlrd and retry."
                ) from exc

            try:
                workbook = xlrd.open_workbook(file_contents=content)
            except Exception as exc:
                raise ValueError(
                    "Could not parse XLS file. Please upload a valid .xls file."
                ) from exc

            if workbook.nsheets == 0:
                raise ValueError("Excel file must include a header row.")
            sheet = workbook.sheet_by_index(0)
            if sheet.nrows < 1:
                raise ValueError("Excel file must include a header row.")
            header_values = sheet.row_values(0)
            headers = [str(h).strip() if h is not None else "" for h in header_values]
            headers = [h for h in headers if h]
            if not headers:
                raise ValueError("Excel file must include a header row.")
            xls_rows: list[dict[str, str]] = []
            total_rows = 0
            for row_index in range(1, sheet.nrows):
                row_cells = sheet.row(row_index)
                if not any(
                    cell.value is not None and str(cell.value).strip()
                    for cell in row_cells
                ):
                    continue
                total_rows += 1
                xls_row: dict[str, str] = {}
                for i, header in enumerate(headers):
                    cell = row_cells[i] if i < len(row_cells) else None
                    if cell is None or cell.value is None:
                        xls_row[header] = ""
                    elif cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            dt = xlrd.xldate_as_datetime(cell.value, workbook.datemode)
                            xls_row[header] = dt.strftime("%Y-%m-%d")
                        except Exception:
                            xls_row[header] = str(cell.value)
                    else:
                        xls_row[header] = str(cell.value)
                if sample_limit is None or len(xls_rows) < sample_limit:
                    xls_rows.append(xls_row)
            return headers, xls_rows, total_rows

        raise ValueError(f"Supported statement files: {spreadsheet_formats_label()}.")

    @staticmethod
    def _parse_column_map(form) -> dict[str, str]:
        pattern = re.compile(r"^column_map\[(.+)\]$")
        mapping: dict[str, str] = {}
        for key, value in form.items():
            match = pattern.match(key)
            if not match:
                continue
            source_col = match.group(1)
            target = str(value).strip() if value is not None else ""
            if not target:
                continue
            mapping[source_col] = target
        return mapping

    @staticmethod
    def _map_rows_with_column_map(
        source_rows: list[dict[str, str]],
        column_map: dict[str, str],
    ) -> list[dict]:
        mapped_rows: list[dict] = []
        for idx, source in enumerate(source_rows, start=1):
            row: dict[str, Any] = {"line_number": idx}
            for source_col, target_field in column_map.items():
                value = source.get(source_col, "")
                row[target_field] = "" if value is None else str(value)
            if any(
                str(v).strip()
                for k, v in row.items()
                if k != "line_number" and v is not None
            ):
                mapped_rows.append(row)
        return mapped_rows

    @staticmethod
    def _parse_manual_lines(form) -> tuple[list[dict], list[str]]:
        pattern = re.compile(r"^lines\[(\d+)\]\[(.+)\]$")
        lines: dict[int, dict] = {}
        errors: list[str] = []

        for key, value in form.items():
            match = pattern.match(key)
            if not match:
                continue
            line_index = int(match.group(1))
            field = match.group(2)
            lines.setdefault(line_index, {})[field] = value

        if not lines:
            return [], []

        results: list[dict] = []
        for idx in sorted(lines):
            data = lines[idx]
            # Skip rows that are entirely empty.
            if not any(str(v).strip() for v in data.values() if v is not None):
                continue

            line_number = data.get("line_number") or idx
            result = {
                "line_number": int(line_number),
                "transaction_date": data.get("transaction_date"),
                "transaction_type": data.get("transaction_type"),
                "amount": data.get("amount"),
                "debit": data.get("debit"),
                "credit": data.get("credit"),
                "description": data.get("description"),
                "reference": data.get("reference"),
                "payee_payer": data.get("payee_payer"),
                "bank_reference": data.get("bank_reference"),
                "check_number": data.get("check_number"),
                "bank_category": data.get("bank_category"),
                "bank_code": data.get("bank_code"),
                "value_date": data.get("value_date"),
                "running_balance": data.get("running_balance"),
                "transaction_id": data.get("transaction_id"),
            }
            results.append(result)

        if not results:
            errors.append("Please add at least one transaction line.")
        return results, errors

    @staticmethod
    def _normalize_mapped_lines(
        lines: list[dict],
        *,
        org_date_fmt: str | None = None,
    ) -> list[dict]:
        normalized: list[dict] = []
        date_fields = ("transaction_date", "value_date")
        decimal_fields = ("amount", "debit", "credit", "running_balance")
        for line in lines:
            row = dict(line)
            tx_type = row.get("transaction_type")
            if tx_type is not None:
                cleaned = str(tx_type).strip().lower()
                row["transaction_type"] = cleaned or None

            for field in date_fields:
                raw = row.get(field)
                if raw is None:
                    continue
                cleaned = str(raw).strip()
                if not cleaned:
                    row[field] = None
                    continue
                parsed_date = _parse_date(cleaned, format=org_date_fmt)
                row[field] = parsed_date if parsed_date is not None else cleaned

            for field in decimal_fields:
                raw = row.get(field)
                if raw is None:
                    continue
                cleaned = str(raw).strip()
                if not cleaned:
                    row[field] = None
                    continue
                parsed_decimal = _parse_decimal(cleaned)
                row[field] = parsed_decimal if parsed_decimal is not None else cleaned

            normalized.append(row)
        return normalized

    @staticmethod
    def _resolve_org_date_format(
        db: Session, organization_id: UUID | None
    ) -> str | None:
        if not organization_id:
            return None
        from app.models.finance.core_org.organization import Organization
        from app.services.formatting_context import DATE_FORMAT_MAP

        org = db.get(Organization, organization_id)
        if not org or not getattr(org, "date_format", None):
            return None
        date_format_key = org.date_format
        if not isinstance(date_format_key, str):
            return None
        date_format = DATE_FORMAT_MAP.get(date_format_key)
        return str(date_format) if date_format is not None else None

    @staticmethod
    def _format_validation_errors(exc: ValidationError) -> list[str]:
        errors: list[str] = []
        for err in exc.errors():
            loc = " -> ".join(str(item) for item in err.get("loc", []))
            msg = err.get("msg", "Invalid value")
            if loc:
                errors.append(f"{loc}: {msg}")
            else:
                errors.append(msg)
        return errors

    def statement_detail_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        *,
        page: int = 1,
        limit: int = 50,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Bank Statement", "banking", db=db)
        context.update(
            self.statement_detail_context(
                db,
                str(auth.organization_id),
                statement_id,
                page=page,
                limit=limit,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/statement_detail.html", context
        )

    def apply_rules_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
    ) -> RedirectResponse:
        """Handle POST to apply categorization rules to a statement."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        service = TransactionCategorizationService()
        result = service.apply_rules_to_statement(db, org_id, coerce_uuid(statement_id))
        db.flush()
        db.commit()

        msg = (
            f"{result.categorized_count}+suggested,"
            f"{result.high_confidence_count}+auto-applied,"
            f"{result.no_match_count}+no+match"
        )
        return RedirectResponse(
            url=f"/finance/banking/statements/{statement_id}?success={msg}",
            status_code=303,
        )

    def accept_suggestion_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        line_id: str,
    ) -> RedirectResponse:
        """Handle POST to accept a categorization suggestion."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        service = TransactionCategorizationService()
        try:
            service.accept_suggestion(
                db, org_id, coerce_uuid(line_id), accepted_by=auth.person_id
            )
            db.flush()
            db.commit()
        except ValueError as exc:
            logger.warning("Accept suggestion failed: %s", exc)
            return RedirectResponse(
                url=f"/finance/banking/statements/{statement_id}?error={exc}",
                status_code=303,
            )

        return RedirectResponse(
            url=f"/finance/banking/statements/{statement_id}?success=Suggestion+accepted",
            status_code=303,
        )

    def reject_suggestion_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        line_id: str,
    ) -> RedirectResponse:
        """Handle POST to reject a categorization suggestion."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        service = TransactionCategorizationService()
        try:
            service.reject_suggestion(db, org_id, coerce_uuid(line_id))
            db.flush()
            db.commit()
        except ValueError as exc:
            logger.warning("Reject suggestion failed: %s", exc)
            return RedirectResponse(
                url=f"/finance/banking/statements/{statement_id}?error={exc}",
                status_code=303,
            )

        return RedirectResponse(
            url=f"/finance/banking/statements/{statement_id}?success=Suggestion+rejected",
            status_code=303,
        )

    # ------------------------------------------------------------------
    # Flat-view response methods (operate on lines across all statements)
    # ------------------------------------------------------------------
    def apply_rules_flat_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        account_id: str,
    ) -> RedirectResponse:
        """Apply categorization rules to all unprocessed lines for a bank account."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        service = TransactionCategorizationService()
        result = service.apply_rules_to_account(db, org_id, coerce_uuid(account_id))
        db.flush()
        db.commit()

        msg = (
            f"{result.categorized_count}+suggested,"
            f"{result.high_confidence_count}+auto-applied,"
            f"{result.no_match_count}+no+match"
        )
        return RedirectResponse(
            url=f"/finance/banking/statements?account_id={account_id}&success={msg}",
            status_code=303,
        )

    def accept_suggestion_flat_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        line_id: str,
    ) -> RedirectResponse:
        """Accept a categorization suggestion from the flat lines view."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        # Resolve the line's bank account for redirect
        line = (
            db.execute(
                select(BankStatementLine)
                .join(BankStatement)
                .where(
                    BankStatementLine.line_id == coerce_uuid(line_id),
                    BankStatement.organization_id == org_id,
                )
            )
            .scalars()
            .first()
        )
        account_id = ""
        if line:
            stmt = db.get(BankStatement, line.statement_id)
            if stmt:
                account_id = str(stmt.bank_account_id)

        service = TransactionCategorizationService()
        try:
            service.accept_suggestion(
                db, org_id, coerce_uuid(line_id), accepted_by=auth.person_id
            )
            db.flush()
            db.commit()
        except ValueError as exc:
            logger.warning("Accept suggestion (flat) failed: %s", exc)
            return RedirectResponse(
                url=f"/finance/banking/statements?account_id={account_id}&error={exc}",
                status_code=303,
            )

        return RedirectResponse(
            url=(
                f"/finance/banking/statements?account_id={account_id}"
                "&success=Suggestion+accepted"
            ),
            status_code=303,
        )

    def reject_suggestion_flat_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        line_id: str,
    ) -> RedirectResponse:
        """Reject a categorization suggestion from the flat lines view."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        # Resolve the line's bank account for redirect
        line = (
            db.execute(
                select(BankStatementLine)
                .join(BankStatement)
                .where(
                    BankStatementLine.line_id == coerce_uuid(line_id),
                    BankStatement.organization_id == org_id,
                )
            )
            .scalars()
            .first()
        )
        account_id = ""
        if line:
            stmt = db.get(BankStatement, line.statement_id)
            if stmt:
                account_id = str(stmt.bank_account_id)

        service = TransactionCategorizationService()
        try:
            service.reject_suggestion(db, org_id, coerce_uuid(line_id))
            db.flush()
            db.commit()
        except ValueError as exc:
            logger.warning("Reject suggestion (flat) failed: %s", exc)
            return RedirectResponse(
                url=f"/finance/banking/statements?account_id={account_id}&error={exc}",
                status_code=303,
            )

        return RedirectResponse(
            url=(
                f"/finance/banking/statements?account_id={account_id}"
                "&success=Suggestion+rejected"
            ),
            status_code=303,
        )

    def delete_statement_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
    ) -> RedirectResponse:
        """Handle POST to delete a bank statement batch."""
        try:
            deleted = bank_statement_service.delete(
                db,
                coerce_uuid(auth.organization_id),
                coerce_uuid(statement_id),
            )
            if not deleted:
                return RedirectResponse(
                    url=f"/finance/banking/statements/{statement_id}?error=Statement+not+found",
                    status_code=303,
                )
            db.flush()
            db.commit()
        except ValueError as exc:
            logger.warning("Delete statement failed: %s", exc)
            return RedirectResponse(
                url=f"/finance/banking/statements/{statement_id}?error={exc}",
                status_code=303,
            )

        return RedirectResponse(
            url="/finance/banking/statements/imports?success=Statement+deleted",
            status_code=303,
        )

    def statement_auto_match_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
    ) -> RedirectResponse:
        """Run auto-match on a specific statement's unmatched lines.

        Delegates to ``AutoReconciliationService.auto_match_statement()``
        for deterministic PaymentIntent + Splynx payment matching.
        """
        from app.services.finance.banking.auto_reconciliation import (
            AutoReconciliationService,
        )

        org_id = coerce_uuid(auth.organization_id)
        stmt_id = coerce_uuid(statement_id)

        auto_svc = AutoReconciliationService()
        result = auto_svc.auto_match_statement(db, org_id, stmt_id)

        if result.matched > 0:
            db.flush()
            db.commit()
            msg = f"Auto-matched+{result.matched}+lines"
            if result.skipped > 0:
                msg += f"+({result.skipped}+skipped)"
            return RedirectResponse(
                url=f"/finance/banking/statements/{statement_id}?success={msg}",
                status_code=303,
            )

        if result.errors:
            error_msg = "Auto-match+errors:+" + ",+".join(result.errors[:3])
            return RedirectResponse(
                url=f"/finance/banking/statements/{statement_id}?error={error_msg}",
                status_code=303,
            )

        return RedirectResponse(
            url=f"/finance/banking/statements/{statement_id}?info=No+new+matches+found",
            status_code=303,
        )

    async def match_statement_line_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        line_id: str,
    ) -> Response:
        """Accept a GL transaction match for a statement line (JSON from Alpine.js)."""
        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        body = await request.json()
        journal_line_id = body.get("journal_line_id")
        force_match = bool(body.get("force_match", False))
        if not journal_line_id:
            return JSONResponse(
                content={"detail": "journal_line_id is required"}, status_code=400
            )

        svc = BankReconciliationService()
        user_id = getattr(auth, "user_id", None) or getattr(auth, "person_id", None)

        try:
            stmt_line = svc.match_statement_line(
                db=db,
                organization_id=org_id,
                statement_line_id=UUID(line_id),
                journal_line_id=UUID(str(journal_line_id)),
                matched_by=user_id,
                force_match=force_match,
            )
            db.flush()
            db.commit()
        except HTTPException:
            raise
        except (ValueError, RuntimeError) as e:
            logger.warning("Statement line match failed: %s", e)
            return JSONResponse(content={"detail": str(e)}, status_code=400)

        # Return updated counters so frontend can update stats without reload
        statement = stmt_line.statement
        matched = statement.matched_lines or 0
        total = statement.total_lines or (
            (statement.matched_lines or 0) + (statement.unmatched_lines or 0)
        )
        match_pct = round(matched / total * 100) if total else 0

        # Resolve source URL + match detail for the matched GL line
        from app.services.finance.banking.bank_reconciliation import (
            _build_source_url,
        )

        source_url = ""
        match_detail: dict[str, str] | None = None
        gl_line = db.get(JournalEntryLine, UUID(str(journal_line_id)))
        if gl_line:
            entry = getattr(gl_line, "journal_entry", None) or getattr(
                gl_line, "entry", None
            )
            if entry:
                source_url = _build_source_url(
                    getattr(entry, "source_document_type", None),
                    getattr(entry, "source_document_id", None),
                    getattr(entry, "entry_id", None),
                )
                match_detail = _build_match_detail(db, entry, source_url)

        return JSONResponse(
            content={
                "status": "ok",
                "matched_lines": matched,
                "unmatched_lines": statement.unmatched_lines or 0,
                "match_pct": match_pct,
                "total_lines": total,
                "source_url": source_url,
                "match_detail": match_detail,
            },
            status_code=200,
        )

    async def batch_match_statement_lines_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
    ) -> Response:
        """Match multiple statement lines to GL entries in a single request."""
        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
            _build_source_url,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        body = await request.json()
        matches: list[dict[str, str]] = body.get("matches", [])
        if not matches:
            return JSONResponse(
                content={"detail": "matches array is required"}, status_code=400
            )

        svc = BankReconciliationService()
        user_id = getattr(auth, "user_id", None) or getattr(auth, "person_id", None)

        results: list[dict[str, object]] = []
        matched_count = 0
        error_count = 0

        for match in matches:
            line_id = match.get("line_id", "")
            journal_line_id = match.get("journal_line_id", "")
            force = bool(match.get("force_match", False))

            if not line_id or not journal_line_id:
                results.append(
                    {"line_id": line_id, "status": "error", "detail": "missing IDs"}
                )
                error_count += 1
                continue

            try:
                svc.match_statement_line(
                    db=db,
                    organization_id=org_id,
                    statement_line_id=UUID(line_id),
                    journal_line_id=UUID(str(journal_line_id)),
                    matched_by=user_id,
                    force_match=force,
                )

                # Resolve source URL + match detail
                source_url = ""
                batch_match_detail: dict[str, str] | None = None
                gl_line = db.get(JournalEntryLine, UUID(str(journal_line_id)))
                if gl_line:
                    entry = getattr(gl_line, "journal_entry", None) or getattr(
                        gl_line, "entry", None
                    )
                    if entry:
                        source_url = _build_source_url(
                            getattr(entry, "source_document_type", None),
                            getattr(entry, "source_document_id", None),
                            getattr(entry, "entry_id", None),
                        )
                        batch_match_detail = _build_match_detail(db, entry, source_url)

                results.append(
                    {
                        "line_id": line_id,
                        "status": "ok",
                        "source_url": source_url,
                        "match_detail": batch_match_detail,
                    }
                )
                matched_count += 1
            except HTTPException as e:
                results.append(
                    {"line_id": line_id, "status": "error", "detail": e.detail}
                )
                error_count += 1
            except (ValueError, RuntimeError) as e:
                logger.warning("Batch match failed for line %s: %s", line_id, e)
                results.append(
                    {"line_id": line_id, "status": "error", "detail": str(e)}
                )
                error_count += 1

        # Commit all successful matches in one transaction
        if matched_count > 0:
            db.flush()
            db.commit()

        # Get final statement counters
        statement = db.get(BankStatement, UUID(statement_id))
        matched = (statement.matched_lines or 0) if statement else 0
        total = (statement.total_lines or 0) if statement else 0
        match_pct = round(matched / total * 100) if total else 0

        return JSONResponse(
            content={
                "status": "ok",
                "matched_count": matched_count,
                "error_count": error_count,
                "results": results,
                "matched_lines": matched,
                "unmatched_lines": (statement.unmatched_lines or 0) if statement else 0,
                "match_pct": match_pct,
                "total_lines": total,
            },
            status_code=200,
        )

    async def unmatch_statement_line_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        line_id: str,
    ) -> Response:
        """Remove a direct match from a statement line (JSON from Alpine.js)."""
        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        svc = BankReconciliationService()

        try:
            stmt_line = svc.unmatch_statement_line(
                db=db,
                organization_id=org_id,
                statement_line_id=UUID(line_id),
            )
            db.flush()
            db.commit()
        except HTTPException:
            raise
        except (ValueError, RuntimeError) as e:
            logger.warning("Statement line unmatch failed: %s", e)
            return JSONResponse(content={"detail": str(e)}, status_code=400)

        # Return updated counters
        statement = stmt_line.statement
        matched = statement.matched_lines or 0
        total = statement.total_lines or (
            (statement.matched_lines or 0) + (statement.unmatched_lines or 0)
        )
        match_pct = round(matched / total * 100) if total else 0

        return JSONResponse(
            content={
                "status": "ok",
                "matched_lines": matched,
                "unmatched_lines": statement.unmatched_lines or 0,
                "match_pct": match_pct,
                "total_lines": total,
            },
            status_code=200,
        )

    async def create_journal_and_match_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        line_id: str,
    ) -> Response:
        """Create a GL journal and match it to a bank line (JSON from Alpine.js).

        Accepts: {counterparty_account_id: str, description?: str}
        """
        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
            _build_source_url,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        body = await request.json()
        counterparty_account_id = body.get("counterparty_account_id")
        description = body.get("description") or None
        if not counterparty_account_id:
            return JSONResponse(
                content={"detail": "counterparty_account_id is required"},
                status_code=400,
            )

        svc = BankReconciliationService()
        user_id = getattr(auth, "user_id", None) or getattr(auth, "person_id", None)

        try:
            stmt_line = svc.create_journal_and_match(
                db=db,
                organization_id=org_id,
                statement_line_id=UUID(line_id),
                counterparty_account_id=UUID(str(counterparty_account_id)),
                description=description,
                matched_by=user_id,
            )
            db.flush()
            db.commit()
        except HTTPException:
            raise
        except (ValueError, RuntimeError) as e:
            logger.warning("Create journal & match failed: %s", e)
            return JSONResponse(content={"detail": str(e)}, status_code=400)

        # Return updated counters
        statement = stmt_line.statement
        matched = statement.matched_lines or 0
        total = statement.total_lines or (
            (statement.matched_lines or 0) + (statement.unmatched_lines or 0)
        )
        match_pct = round(matched / total * 100) if total else 0

        # Resolve source URL for the newly matched GL line
        source_url = ""
        match_detail: dict[str, str] | None = None
        gl_line = db.get(JournalEntryLine, stmt_line.matched_journal_line_id)
        if gl_line:
            entry = getattr(gl_line, "journal_entry", None) or getattr(
                gl_line, "entry", None
            )
            if entry:
                source_url = _build_source_url(
                    getattr(entry, "source_document_type", None),
                    getattr(entry, "source_document_id", None),
                    getattr(entry, "entry_id", None),
                )
                match_detail = _build_match_detail(db, entry, source_url)

        return JSONResponse(
            content={
                "status": "ok",
                "matched_lines": matched,
                "unmatched_lines": statement.unmatched_lines or 0,
                "match_pct": match_pct,
                "total_lines": total,
                "source_url": source_url,
                "match_detail": match_detail,
            },
            status_code=200,
        )

    def scored_candidates_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        line_id: str,
        *,
        date_from: str | None = None,
        date_to: str | None = None,
        source_type: str | None = None,
        search: str | None = None,
        direction: str | None = None,
        hide_matched: bool = False,
        sort: str = "relevance",
        page: int = 1,
        per_page: int = 25,
    ) -> Response:
        """Return scored GL candidates for a specific statement line (JSON)."""
        from datetime import date as date_type

        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        # Parse date strings to date objects
        parsed_date_from: date_type | None = None
        parsed_date_to: date_type | None = None
        if date_from:
            try:
                parsed_date_from = date_type.fromisoformat(date_from)
            except ValueError:
                pass
        if date_to:
            try:
                parsed_date_to = date_type.fromisoformat(date_to)
            except ValueError:
                pass

        svc = BankReconciliationService()
        result = svc.get_scored_candidates_for_line(
            db=db,
            organization_id=org_id,
            statement_id=UUID(statement_id),
            statement_line_id=UUID(line_id),
            date_from=parsed_date_from,
            date_to=parsed_date_to,
            source_type=source_type or None,
            search=search or None,
            direction=direction or None,
            hide_matched=hide_matched,
            sort=sort,
            page=max(1, page),
            per_page=max(1, min(per_page, 100)),
        )

        return JSONResponse(content=result, status_code=200)

    async def multi_match_statement_line_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        line_id: str,
    ) -> Response:
        """Match one bank line to multiple GL entries (JSON from Alpine.js)."""
        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
            _build_source_url,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        body = await request.json()
        journal_line_ids_raw = body.get("journal_line_ids", [])
        force_match = bool(body.get("force_match", False))

        if not journal_line_ids_raw:
            return JSONResponse(
                content={"detail": "journal_line_ids is required"},
                status_code=400,
            )

        journal_line_ids = [UUID(str(jl_id)) for jl_id in journal_line_ids_raw]

        svc = BankReconciliationService()
        user_id = getattr(auth, "user_id", None) or getattr(auth, "person_id", None)

        try:
            stmt_line = svc.multi_match_statement_line(
                db=db,
                organization_id=org_id,
                statement_line_id=UUID(line_id),
                journal_line_ids=journal_line_ids,
                matched_by=user_id,
                force_match=force_match,
            )
            db.flush()
            db.commit()
        except HTTPException:
            raise
        except (ValueError, RuntimeError) as e:
            logger.warning("Multi-match failed: %s", e)
            return JSONResponse(content={"detail": str(e)}, status_code=400)

        # Return updated counters + source URLs for all matched GL lines
        statement = stmt_line.statement
        matched = statement.matched_lines or 0
        total = statement.total_lines or (
            (statement.matched_lines or 0) + (statement.unmatched_lines or 0)
        )
        match_pct = round(matched / total * 100) if total else 0

        # Resolve source URL + match detail for the primary (first) GL line
        source_url = ""
        multi_match_detail: dict[str, str] | None = None
        if journal_line_ids:
            gl_line = db.get(JournalEntryLine, journal_line_ids[0])
            if gl_line:
                entry = getattr(gl_line, "journal_entry", None) or getattr(
                    gl_line, "entry", None
                )
                if entry:
                    source_url = _build_source_url(
                        getattr(entry, "source_document_type", None),
                        getattr(entry, "source_document_id", None),
                        getattr(entry, "entry_id", None),
                    )
                    multi_match_detail = _build_match_detail(db, entry, source_url)

        return JSONResponse(
            content={
                "status": "ok",
                "matched_lines": matched,
                "unmatched_lines": statement.unmatched_lines or 0,
                "match_pct": match_pct,
                "total_lines": total,
                "source_url": source_url,
                "match_detail": multi_match_detail,
                "match_count": len(journal_line_ids),
            },
            status_code=200,
        )

    async def bulk_delete_statements_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ):
        """Handle bulk delete statements request."""
        from app.schemas.bulk_actions import BulkActionRequest
        from app.services.finance.banking.bulk import get_statement_bulk_service

        body = await request.json()
        req = BulkActionRequest(**body)
        service = get_statement_bulk_service(
            db,
            coerce_uuid(auth.organization_id),
            coerce_uuid(auth.user_id),
        )
        return await service.bulk_delete(req.ids)

    async def bulk_export_statements_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ):
        """Handle bulk export statements request."""
        from app.schemas.bulk_actions import BulkExportRequest
        from app.services.finance.banking.bulk import get_statement_bulk_service

        body = await request.json()
        req = BulkExportRequest(**body)
        service = get_statement_bulk_service(
            db,
            coerce_uuid(auth.organization_id),
            coerce_uuid(auth.user_id),
        )
        return await service.bulk_export(req.ids, req.format)
