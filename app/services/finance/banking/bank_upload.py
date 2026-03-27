"""
Bank Upload Service.

Generates bank upload files for bulk payments.
Supports multiple bank formats (Zenith, Access, GTBank, etc.).
Reusable across payroll, AP bills, and any bulk payment scenarios.
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from typing import Literal

from sqlalchemy.orm import Session

from app.services.finance.banking.bank_directory import BankDirectoryService

logger = logging.getLogger(__name__)

BankFormat = Literal["zenith", "access", "gtbank", "generic"]


@dataclass
class PaymentItem:
    """
    Generic payment item for bank upload generation.

    Reusable across payroll, AP bills, and other bulk payments.
    """

    reference: str  # Transaction reference
    beneficiary_name: str  # Recipient name
    amount: Decimal  # Payment amount
    account_number: str  # Beneficiary account number
    bank_name: str  # Beneficiary bank name (will be resolved to code)
    bank_code: str | None = None  # Beneficiary bank code (if known)
    beneficiary_code: str | None = None  # Internal code (employee ID, supplier code)
    narration: str | None = None  # Payment narration/memo


@dataclass
class BankUploadResult:
    """Result of bank upload file generation."""

    content: bytes
    filename: str
    content_type: str
    row_count: int
    total_amount: Decimal
    errors: list[str]


class BankUploadService:
    """
    Service for generating bank upload files.

    Supports multiple bank formats and automatically resolves
    bank codes from bank names using the bank directory.
    """

    def __init__(self, db: Session):
        self.db = db
        self.bank_directory = BankDirectoryService(db)

    def generate_upload(
        self,
        items: list[PaymentItem],
        source_account_number: str,
        payment_date: date,
        bank_format: BankFormat = "zenith",
        batch_reference: str | None = None,
    ) -> BankUploadResult:
        """
        Generate bank upload file for bulk payments.

        Args:
            items: List of payment items
            source_account_number: Debit account number
            payment_date: Payment date
            bank_format: Target bank format
            batch_reference: Optional batch reference prefix

        Returns:
            BankUploadResult with file content and metadata
        """
        if bank_format == "zenith":
            return self._generate_zenith_format(
                items, source_account_number, payment_date, batch_reference
            )
        elif bank_format == "access":
            return self._generate_access_format(
                items, source_account_number, payment_date, batch_reference
            )
        elif bank_format == "gtbank":
            return self._generate_gtbank_format(
                items, source_account_number, payment_date, batch_reference
            )
        else:
            return self._generate_generic_format(
                items, source_account_number, payment_date, batch_reference
            )

    def _resolve_bank_code(self, item: PaymentItem) -> str:
        """
        Resolve bank code for a payment item.

        Uses provided bank_code if available, otherwise looks up from bank_name.
        Returns bank code as a string, zero-padded to 3 digits.
        """
        code = item.bank_code
        if not code:
            code = self.bank_directory.lookup_bank_code(item.bank_name)

        if not code:
            return ""

        # Ensure bank code is formatted as 3-digit string (e.g., "044", "057")
        code_str = str(code).strip()
        if code_str.isdigit():
            return code_str.zfill(3)
        return code_str

    def _format_account_number(self, account_number: str) -> str:
        """
        Format account number as 10-digit string with leading zeros preserved.

        Nigerian bank accounts are 10 digits (NUBAN format).
        """
        if not account_number:
            return ""

        # Remove any spaces or dashes
        cleaned = str(account_number).strip().replace(" ", "").replace("-", "")

        # If numeric, zero-pad to 10 digits
        if cleaned.isdigit():
            return cleaned.zfill(10)

        return cleaned

    def _generate_zenith_format(
        self,
        items: list[PaymentItem],
        source_account_number: str,
        payment_date: date,
        batch_reference: str | None = None,
    ) -> BankUploadResult:
        """
        Generate Zenith Bank Corporate I-Bank bulk payment Excel file.

        Matches the official Zenith template exactly:
        - Column headers are the full descriptive text from the bank template
        - Text columns use @ (text) format to preserve leading zeros
        - Amount uses 0.00 number format
        - Date is a datetime with dd/mm/yyyy format
        """
        from datetime import datetime

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Exact headers from official Zenith Corporate I-Bank template
        headers_and_formats = [
            (
                "TRANSACTION REFERENCE NUMBER (MANDATORY FIELD) This is a unique "
                "reference created by the payer and used to identify a payment. "
                "Must not contain commas semi-colon apostrophe or space. "
                "Text format. Alpha-numeric(max. 30 characters)",
                "@",
            ),
            (
                "BENEFICIARY NAME (MANDATORY FIELD) Text format. "
                "Alpha-numeric(max. 100 characters)",
                "@",
            ),
            (
                "PAYMENT AMOUNT (MANDATORY FIELD) Number format with 2 decimal "
                "digits. Must not contain commas semi-colon apostrophe or spaces",
                "0.00;[Red]0.00",
            ),
            (
                "PAYMENT DUE DATE (MANDATORY FIELD) This is the effective date "
                "of payment. Format is DD/MM/YYYY (max. 10 characters)",
                "dd/mm/yyyy;@",
            ),
            (
                "BENEFICIARY CODE (MANDATORY FIELD) Unique code assigned by "
                "Payer to the beneficiary. Used on Corporate I-Bank to search "
                "for payments made to the beneficiary. Alphanumeric e.g. staff "
                "number. RC no. or name (max. 35 characters)",
                "@",
            ),
            (
                "BENEFICIARY ACCOUNT NUMBER (MANDATORY FIELD) Numeric (10 digits)",
                "@",
            ),
            (
                "BENEFICIARY BANK SORT CODE (MANDATORY FIELD) This is used to "
                "represent Beneficiary Bank Name and Payment routing method. "
                "Leave blank for Zenith beneficiaries or use 057. Use first "
                "3-digits for Instant transfer via InterSwitch. Use 9-digits "
                "for non-instant transfer via NEFT",
                "@",
            ),
            (
                "DEBIT ACCOUNT NUMBER (MANDATORY FIELD) This is the account "
                "number to debit. Number format (max. 10 digits)",
                "@",
            ),
        ]

        # Write header row (no bold — matches Zenith template exactly)
        for col_idx, (header_text, fmt) in enumerate(headers_and_formats, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.number_format = fmt

        errors: list[str] = []
        total_amount = Decimal("0")
        row_count = 0

        formatted_source_account = self._format_account_number(source_account_number)
        payment_datetime = datetime.combine(payment_date, datetime.min.time())

        for item in items:
            bank_code = self._resolve_bank_code(item)
            if not bank_code:
                errors.append(
                    f"Bank code not found for: {item.beneficiary_name} ({item.bank_name})"
                )

            account_number = self._format_account_number(item.account_number)
            row_count += 1
            row = row_count + 1  # +1 for header

            # Col 1: Transaction reference (text, no commas/semicolons/spaces)
            ref = (item.reference or "").replace(",", "").replace(";", "")
            c = ws.cell(row=row, column=1, value=ref)
            c.number_format = "@"

            # Col 2: Beneficiary name (text)
            c = ws.cell(row=row, column=2, value=item.beneficiary_name)
            c.number_format = "@"

            # Col 3: Amount (number, 2 decimals)
            c = ws.cell(row=row, column=3, value=float(item.amount))
            c.number_format = "0.00;[Red]0.00"

            # Col 4: Payment date (datetime)
            c = ws.cell(row=row, column=4, value=payment_datetime)
            c.number_format = "dd/mm/yyyy;@"

            # Col 5: Beneficiary code (text, mandatory)
            c = ws.cell(
                row=row, column=5, value=item.beneficiary_code or item.reference
            )
            c.number_format = "@"

            # Col 6: Account number (text, 10 digits)
            c = ws.cell(row=row, column=6, value=account_number)
            c.number_format = "@"

            # Col 7: Sort code (text)
            c = ws.cell(row=row, column=7, value=bank_code)
            c.number_format = "@"

            # Col 8: Debit account (text)
            c = ws.cell(row=row, column=8, value=formatted_source_account)
            c.number_format = "@"

            total_amount += item.amount

        buffer = io.BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()
        filename = f"bank_upload_zenith_{payment_date.strftime('%Y%m%d')}.xlsx"

        return BankUploadResult(
            content=content,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            row_count=row_count,
            total_amount=total_amount,
            errors=errors,
        )

    def _generate_access_format(
        self,
        items: list[PaymentItem],
        source_account_number: str,
        payment_date: date,
        batch_reference: str | None = None,
    ) -> BankUploadResult:
        """
        Generate Access Bank upload format.

        Similar to Zenith but with slightly different column order.
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow(
            [
                "Serial No",
                "Beneficiary Account Number",
                "Beneficiary Bank Code",
                "Beneficiary Name",
                "Amount",
                "Narration",
            ]
        )

        errors: list[str] = []
        total_amount = Decimal("0")
        row_count = 0

        for idx, item in enumerate(items, start=1):
            bank_code = self._resolve_bank_code(item)
            if not bank_code:
                errors.append(
                    f"Bank code not found for: {item.beneficiary_name} ({item.bank_name})"
                )

            narration = item.narration or f"Payment to {item.beneficiary_name}"

            writer.writerow(
                [
                    idx,
                    self._format_account_number(item.account_number),
                    bank_code,
                    item.beneficiary_name,
                    str(item.amount),
                    narration,
                ]
            )
            total_amount += item.amount
            row_count += 1

        content = output.getvalue().encode("utf-8")
        filename = f"bank_upload_access_{payment_date.strftime('%Y%m%d')}.csv"

        return BankUploadResult(
            content=content,
            filename=filename,
            content_type="text/csv",
            row_count=row_count,
            total_amount=total_amount,
            errors=errors,
        )

    def _generate_gtbank_format(
        self,
        items: list[PaymentItem],
        source_account_number: str,
        payment_date: date,
        batch_reference: str | None = None,
    ) -> BankUploadResult:
        """
        Generate GTBank upload format.
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow(
            [
                "Account Number",
                "Bank Code",
                "Amount",
                "Beneficiary Name",
                "Remarks",
            ]
        )

        errors: list[str] = []
        total_amount = Decimal("0")
        row_count = 0

        for item in items:
            bank_code = self._resolve_bank_code(item)
            if not bank_code:
                errors.append(
                    f"Bank code not found for: {item.beneficiary_name} ({item.bank_name})"
                )

            remarks = item.narration or item.reference

            writer.writerow(
                [
                    self._format_account_number(item.account_number),
                    bank_code,
                    str(item.amount),
                    item.beneficiary_name,
                    remarks,
                ]
            )
            total_amount += item.amount
            row_count += 1

        content = output.getvalue().encode("utf-8")
        filename = f"bank_upload_gtbank_{payment_date.strftime('%Y%m%d')}.csv"

        return BankUploadResult(
            content=content,
            filename=filename,
            content_type="text/csv",
            row_count=row_count,
            total_amount=total_amount,
            errors=errors,
        )

    def _generate_generic_format(
        self,
        items: list[PaymentItem],
        source_account_number: str,
        payment_date: date,
        batch_reference: str | None = None,
    ) -> BankUploadResult:
        """
        Generate generic bank upload format.

        A universal format that can be adapted for any bank.
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow(
            [
                "Reference",
                "Beneficiary Name",
                "Account Number",
                "Bank Code",
                "Bank Name",
                "Amount",
                "Date",
                "Narration",
            ]
        )

        errors: list[str] = []
        total_amount = Decimal("0")
        row_count = 0
        date_str = payment_date.strftime("%Y-%m-%d")

        for item in items:
            bank_code = self._resolve_bank_code(item)
            if not bank_code:
                errors.append(
                    f"Bank code not found for: {item.beneficiary_name} ({item.bank_name})"
                )

            narration = item.narration or f"Payment - {item.reference}"

            writer.writerow(
                [
                    item.reference,
                    item.beneficiary_name,
                    self._format_account_number(item.account_number),
                    bank_code,
                    item.bank_name,
                    str(item.amount),
                    date_str,
                    narration,
                ]
            )
            total_amount += item.amount
            row_count += 1

        content = output.getvalue().encode("utf-8")
        filename = f"bank_upload_{payment_date.strftime('%Y%m%d')}.csv"

        return BankUploadResult(
            content=content,
            filename=filename,
            content_type="text/csv",
            row_count=row_count,
            total_amount=total_amount,
            errors=errors,
        )


def bank_upload_service(db: Session) -> BankUploadService:
    """Create a BankUploadService instance."""
    return BankUploadService(db)
